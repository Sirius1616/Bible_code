#!/usr/bin/env python
# 30_subheadfinder_auto.py

import os
import re
from openpyxl import load_workbook
import csv

def clean_text_content(text):
    """Clean text while preserving apostrophes and quotation marks"""
    if not text:
        return ""
    
    # PRESERVE APOSTROPHES - don't remove them!
    # Keep all types of apostrophes and quotes: ' \u2018 \u2019 \u0027 \u0060 \u00B4 \u02BC \u2032 \u055A \u05F3 \uFF07
    # Only remove truly problematic characters
    text = re.sub(r'[^\w\s\-–—:;,.!?()\'\u2018\u2019\u0027\u0060\u00B4\u02BC\u2032\u055A\u05F3\uFF07]', '', text)
    # Normalize whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def normalize_apostrophes(text):
    """Normalize all types of apostrophes to a standard form for matching"""
    if not text:
        return ""
    
    # Convert all apostrophe variants to standard ASCII apostrophe
    apostrophe_variants = ['\u2018', '\u2019', '\u0027', '\u0060', '\u00B4', '\u02BC', '\u2032', '\u055A', '\u05F3', '\uFF07']
    for variant in apostrophe_variants:
        text = text.replace(variant, "'")
    
    return text

def load_txt_subheads(txt_file_path):
    """Load subhead titles from TXT file and clean them"""
    with open(txt_file_path, 'r', encoding='utf-8') as file:
        subheads = [line.strip() for line in file if line.strip()]
        return [clean_text_content(subhead) for subhead in subheads]

def extract_xlsx_structure(xlsx_file_path):
    """
    Extract the complete structure from XLSX file with proper text cleaning
    """
    wb = load_workbook(xlsx_file_path)
    ws = wb.active
    
    subheads = []  # (cleaned_text, row_number, original_text)
    verses = []    # (reference, verse_text, row_number)
    chapters = []  # (chapter_number, row_number)
    
    # Extract book name from filename (more flexible matching)
    filename = os.path.basename(xlsx_file_path)
    # Look for book name pattern (number followed by text)
    book_match = re.search(r'(?:^|\D)(\d+)[-\s]*([A-Za-z]+)', filename)
    if book_match:
        current_book = book_match.group(2)
    else:
        current_book = "Unknown"  # Fallback using original logic
    
    current_chapter = "1"
    current_verse = None
    verse_text_parts = []
    row_number = 0
    
    for row in ws.iter_rows(values_only=True):
        row_number += 1
        if not row or row[1] is None:
            continue
            
        row_type = str(row[1])
        content = str(row[2]) if len(row) > 2 and row[2] is not None else ""
        
        # SUBHEAD DETECTION
        if "SUBHEAD" in row_type and content:
            cleaned_content = clean_text_content(content)
            if cleaned_content:
                subheads.append((cleaned_content, row_number, content))  # Store both cleaned and original
        
        # CHAPTER DETECTION  
        elif "CHAPTER NUMBERS" in row_type and content.strip():
            # Extract chapter number
            chapter_match = ''.join(c for c in content if c.isdigit())
            if chapter_match:
                current_chapter = chapter_match
                chapters.append((current_chapter, row_number))
        
        # VERSE NUMBER DETECTION
        elif "VERSE NUMBERS" in row_type and content.strip():
            # Save previous verse if exists
            if current_verse is not None and verse_text_parts:
                full_text = " ".join(verse_text_parts).strip()
                if full_text:
                    reference = f"{current_book} {current_chapter}:{current_verse}"
                    verses.append((reference, full_text, row_number))
            
            # Start new verse
            verse_match = ''.join(c for c in content if c.isdigit())
            if verse_match:
                current_verse = verse_match
            verse_text_parts = []
        
        # VERSE TEXT DETECTION
        elif "SCRIPTURE TEXT" in row_type and content and content not in ['""', '']:
            clean_content = content.replace('"', '').replace('""', '').strip()
            if clean_content:
                verse_text_parts.append(clean_content)
    
    # Add the last verse
    if current_verse is not None and verse_text_parts:
        full_text = " ".join(verse_text_parts).strip()
        if full_text:
            reference = f"{current_book} {current_chapter}:{current_verse}"
            verses.append((reference, full_text, row_number))
    
    return subheads, verses, chapters

def find_verse_after_subhead(subhead_row, verses):
    """Find the first verse that comes after a subhead"""
    for ref, text, verse_row in verses:
        if verse_row > subhead_row:
            return ref, text
    return None, None

def match_subheads_to_verses(subhead_phrases, subheads, verses):
    """Match subhead titles to the verses that follow them with fuzzy matching"""
    results = []
    not_found = []
    
    # Create mappings for both cleaned and original subhead text with normalized apostrophes
    subhead_clean_map = {}  # cleaned text -> (row_number, original_text)
    subhead_normalized_map = {}  # normalized text (apostrophes standardized) -> (row_number, original_text)
    subhead_original_map = {}  # original text -> row_number
    
    for cleaned_text, row_num, original_text in subheads:
        subhead_clean_map[cleaned_text] = (row_num, original_text)
        subhead_original_map[original_text] = row_num
        
        # Create normalized version for apostrophe-insensitive matching
        normalized_text = normalize_apostrophes(cleaned_text)
        if normalized_text not in subhead_normalized_map:
            subhead_normalized_map[normalized_text] = []
        subhead_normalized_map[normalized_text].append((row_num, original_text))
    
    for subhead_phrase in subhead_phrases:
        cleaned_phrase = clean_text_content(subhead_phrase)
        normalized_phrase = normalize_apostrophes(cleaned_phrase)
        found = False
        
        # Strategy 1: Try exact match with cleaned text first
        if cleaned_phrase in subhead_clean_map:
            row_num, original_text = subhead_clean_map[cleaned_phrase]
            verse_ref, verse_text = find_verse_after_subhead(row_num, verses)
            
            if verse_ref:
                results.append({
                    'reference': verse_ref,
                    'subhead': subhead_phrase,
                    'original_xlsx_subhead': original_text
                })
                found = True
        
        # Strategy 2: Try normalized apostrophe matching
        if not found and normalized_phrase in subhead_normalized_map:
            matches = subhead_normalized_map[normalized_phrase]
            # Use the first match (should be sufficient for most cases)
            row_num, original_text = matches[0]
            verse_ref, verse_text = find_verse_after_subhead(row_num, verses)
            
            if verse_ref:
                results.append({
                    'reference': verse_ref,
                    'subhead': subhead_phrase,
                    'original_xlsx_subhead': original_text
                })
                found = True
        
        # Strategy 3: If not found, try fuzzy matching with normalized text
        if not found:
            best_match = None
            best_score = 0
            
            for normalized_text, matches in subhead_normalized_map.items():
                # Check if the normalized phrases are similar
                if (normalized_phrase in normalized_text or normalized_text in normalized_phrase):
                    similarity = len(set(normalized_phrase.split()) & set(normalized_text.split())) / max(len(normalized_phrase.split()), len(normalized_text.split()))
                    if similarity > 0.7:  # 70% similarity threshold
                        row_num, original_text = matches[0]
                        verse_ref, verse_text = find_verse_after_subhead(row_num, verses)
                        if verse_ref:
                            results.append({
                                'reference': verse_ref,
                                'subhead': subhead_phrase,
                                'original_xlsx_subhead': original_text
                            })
                            found = True
                            break
            
            # Strategy 4: Try matching with original text (before cleaning)
            if not found and subhead_phrase in subhead_original_map:
                row_num = subhead_original_map[subhead_phrase]
                verse_ref, verse_text = find_verse_after_subhead(row_num, verses)
                if verse_ref:
                    results.append({
                        'reference': verse_ref,
                        'subhead': subhead_phrase,
                        'original_xlsx_subhead': subhead_phrase
                    })
                    found = True
        
        if not found:
            not_found.append(subhead_phrase)
    
    return results, not_found

def find_matching_files():
    """Find ALL matching TXT and XLSX files in the current directory"""
    files = os.listdir('.')
    
    # Find ALL TXT files that start with number pattern (e.g., "01-")
    txt_files = [f for f in files if f.lower().endswith('.txt') and re.search(r'\d+', f)]
    
    # Find ALL XLSX files that start with number pattern
    xlsx_files = [f for f in files if f.lower().endswith(('.xlsx', '.xls')) and re.search(r'\d+', f)]
    
    matches = []
    
    # Match files by their number prefix (e.g., "01-" from "01-Genesis.txt" and "01-Genesis.xlsx")
    for txt_file in txt_files:
        # Extract the number prefix from TXT filename
        txt_number_match = re.search(r'(\d+)', txt_file)
        if txt_number_match:
            txt_number = txt_number_match.group(1)
            
            # Find matching XLSX file with same number
            for xlsx_file in xlsx_files:
                xlsx_number_match = re.search(r'(\d+)', xlsx_file)
                if xlsx_number_match and xlsx_number_match.group(1) == txt_number:
                    matches.append((txt_file, xlsx_file))
                    break
        else:
            # If no number found, try to match by basename without extension
            txt_base = os.path.splitext(txt_file)[0]
            for xlsx_file in xlsx_files:
                xlsx_base = os.path.splitext(xlsx_file)[0]
                if txt_base in xlsx_base or xlsx_base in txt_base:
                    matches.append((txt_file, xlsx_file))
                    break
    
    # If no matches found by numbers, try simple extension-based matching
    if not matches:
        txt_files = [f for f in files if f.lower().endswith('.txt')]
        xlsx_files = [f for f in files if f.lower().endswith(('.xlsx', '.xls'))]
        if len(txt_files) == 1 and len(xlsx_files) == 1:
            matches.append((txt_files[0], xlsx_files[0]))
    
    return matches

def process_single_pair(txt_file, xlsx_file):
    """Process a single pair of files"""
    print(f"Processing: {txt_file} with {xlsx_file}")
    
    # Load data
    subhead_phrases = load_txt_subheads(txt_file)
    print(f"  Loaded {len(subhead_phrases)} subhead titles from TXT")
    
    subheads, verses, chapters = extract_xlsx_structure(xlsx_file)
    print(f"  Found {len(subheads)} subheads, {len(verses)} verses, and {len(chapters)} chapters in XLSX")
    
    # Show samples for verification
    print("  Sample subheads from TXT (with preserved apostrophes):")
    for i, phrase in enumerate(subhead_phrases[:3], 1):
        print(f"    {i}. '{phrase}'")
    
    print("  Sample subheads from XLSX (with preserved apostrophes):")
    for i, (cleaned_text, row_num, original_text) in enumerate(subheads[:3], 1):
        print(f"    {i}. '{original_text}'")
    
    # Match subheads to verses
    results, not_found = match_subheads_to_verses(subhead_phrases, subheads, verses)
    print(f"  Results: {len(results)} matched, {len(not_found)} not found")
    
    # Create output filename (same as TXT file but with .csv extension)
    output_file = os.path.splitext(txt_file)[0] + '.csv'
    
    # Write output
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t')
        
        # Unmatched items first
        if not_found:
            writer.writerow(["ERROR: Could not find verses for these subheads:"])
            writer.writerow([])
            for phrase in not_found:
                writer.writerow([f"NOT_FOUND: {phrase}"])
            writer.writerow([])
            writer.writerow([])
        
        # Matches
        writer.writerow(["Reference", "Subhead"])
        for result in results:
            writer.writerow([result['reference'], result['subhead']])
    
    print(f"  Output saved to: {output_file}")
    
    # Show some examples
    if results:
        print("  First 3 matches:")
        for result in results[:3]:
            print(f"    {result['reference']} -> {result['subhead']}")
    
    # Debug: show some not found items
    if not_found:
        print("  Sample not found items:")
        for phrase in not_found[:5]:
            print(f"    NOT FOUND: '{phrase}'")
    
    return len(results), len(not_found)

def main():
    print("=== Automated Bible Subhead to Verse Matcher ===")
    print("Now processes ALL files, preserves apostrophes, and works with any book")
    
    # Find ALL matching file pairs
    file_pairs = find_matching_files()
    
    if not file_pairs:
        print("No matching files found.")
        print("Looking for files with numbers in their names (e.g., '01-Genesis.txt', '02-Exodus.xlsx')")
        print("Files in current directory:")
        for f in os.listdir('.'):
            print(f"  {f}")
        return
    
    print(f"Found {len(file_pairs)} matching file pair(s):")
    for txt, xlsx in file_pairs:
        print(f"  {txt} -> {xlsx}")
    
    print("\n" + "="*60)
    
    total_matched = 0
    total_not_found = 0
    
    # Process each file pair
    for txt_file, xlsx_file in file_pairs:
        try:
            matched, not_found = process_single_pair(txt_file, xlsx_file)
            total_matched += matched
            total_not_found += not_found
            print()
        except Exception as e:
            print(f"Error processing {txt_file}: {e}")
            import traceback
            traceback.print_exc()
            print()
    
    print("="*60)
    print("Processing complete!")
    print(f"Total across all files: {total_matched} matched, {total_not_found} not found")

if __name__ == "__main__":
    main()