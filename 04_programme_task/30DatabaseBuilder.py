#!/usr/bin/env python
# 30DatabaseBuilder.py

import os
import re
from openpyxl import load_workbook
import csv

def clean_text_content(text):
    """Clean text while preserving apostrophes and quotation marks"""
    if not text:
        return ""
    
    # PRESERVE APOSTROPHES - don't remove them!
    # Keep all types of apostrophes and quotes: ' \u2018 \u2019 \u0027 \u0060 \u00B4 \u02BC \u2032 \u055A \u05F3 \uFF07
    # Only remove truly problematic characters
    text = re.sub(r'[^\w\s\-–—:;,.!?()\'\u2018\u2019\u0027\u0060\u00B4\u02BC\u2032\u055A\u05F3\uFF07]', '', text)
    # Normalize whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def normalize_apostrophes(text):
    """Normalize all types of apostrophes to a standard form for matching"""
    if not text:
        return ""
    
    # Convert all apostrophe variants to standard ASCII apostrophe
    apostrophe_variants = ['\u2018', '\u2019', '\u0027', '\u0060', '\u00B4', '\u02BC', '\u2032', '\u055A', '\u05F3', '\uFF07']
    for variant in apostrophe_variants:
        text = text.replace(variant, "'")
    
    return text

def remove_verse_number(phrase):
    """Remove verse numbers from the beginning of phrases"""
    # Remove numbers and any following space from the beginning
    return re.sub(r'^\d+\s*', '', phrase).strip()

def load_txt_phrases(txt_file_path):
    """Load verse phrases from TXT file, remove verse numbers, and clean them"""
    with open(txt_file_path, 'r', encoding='utf-8') as file:
        phrases = [line.strip() for line in file if line.strip()]
        # Remove verse numbers and clean
        return [clean_text_content(remove_verse_number(phrase)) for phrase in phrases]

def extract_xlsx_structure(xlsx_file_path):
    """
    Extract the complete structure from XLSX file with proper text cleaning
    """
    wb = load_workbook(xlsx_file_path)
    ws = wb.active
    
    verses = []    # (reference, verse_text, row_number)
    chapters = []  # (chapter_number, row_number)
    
    # Extract book name from filename (more flexible matching)
    filename = os.path.basename(xlsx_file_path)
    # Look for book name pattern (number followed by text)
    book_match = re.search(r'(?:^|\D)(\d+)[-\s]*([A-Za-z]+)', filename)
    if book_match:
        current_book = book_match.group(2)
    else:
        current_book = "Genesis"  # Default to Genesis
    
    current_chapter = "1"
    current_verse = None
    verse_text_parts = []
    row_number = 0
    
    for row in ws.iter_rows(values_only=True):
        row_number += 1
        if not row or row[1] is None:
            continue
            
        row_type = str(row[1])
        content = str(row[2]) if len(row) > 2 and row[2] is not None else ""
        
        # CHAPTER DETECTION  
        if "CHAPTER NUMBERS" in row_type and content.strip():
            # Extract chapter number
            chapter_match = ''.join(c for c in content if c.isdigit())
            if chapter_match:
                current_chapter = chapter_match
                chapters.append((current_chapter, row_number))
        
        # VERSE NUMBER DETECTION
        elif "VERSE NUMBERS" in row_type and content.strip():
            # Save previous verse if exists
            if current_verse is not None and verse_text_parts:
                full_text = " ".join(verse_text_parts).strip()
                if full_text:
                    reference = f"{current_book} {current_chapter}:{current_verse}"
                    verses.append((reference, full_text, row_number))
            
            # Start new verse
            verse_match = ''.join(c for c in content if c.isdigit())
            if verse_match:
                current_verse = verse_match
            verse_text_parts = []
        
        # VERSE TEXT DETECTION
        elif "SCRIPTURE TEXT" in row_type and content and content not in ['""', '']:
            clean_content = content.replace('"', '').replace('""', '').strip()
            if clean_content:
                verse_text_parts.append(clean_content)
    
    # Add the last verse
    if current_verse is not None and verse_text_parts:
        full_text = " ".join(verse_text_parts).strip()
        if full_text:
            reference = f"{current_book} {current_chapter}:{current_verse}"
            verses.append((reference, full_text, row_number))
    
    return verses, chapters

def match_phrases_to_verses(phrases, verses):
    """Match phrases to the verses that contain them"""
    results = []
    not_found = []
    
    # Create normalized versions of all verses for better matching
    normalized_verses = []
    for ref, text, row_num in verses:
        normalized_text = normalize_apostrophes(text.lower())
        normalized_verses.append((ref, text, normalized_text, row_num))
    
    for original_phrase in phrases:
        cleaned_phrase = clean_text_content(original_phrase)
        normalized_phrase = normalize_apostrophes(cleaned_phrase.lower())
        found = False
        
        # Try to find which verse contains this phrase
        for ref, original_text, normalized_text, row_num in normalized_verses:
            # Check if phrase is contained in the verse text (case-insensitive)
            if normalized_phrase in normalized_text:
                results.append({
                    'reference': ref,
                    'phrase': original_phrase,
                    'verse_text': original_text
                })
                found = True
                break
        
        # If not found, try fuzzy matching
        if not found:
            best_match = None
            best_score = 0
            
            for ref, original_text, normalized_text, row_num in normalized_verses:
                # Check for significant word overlap
                phrase_words = set(normalized_phrase.split())
                verse_words = set(normalized_text.split())
                common_words = phrase_words.intersection(verse_words)
                
                if len(common_words) / max(len(phrase_words), 1) > 0.6:  # 60% match
                    score = len(common_words) / len(phrase_words)
                    if score > best_score:
                        best_score = score
                        best_match = (ref, original_text)
            
            if best_match and best_score > 0.6:
                ref, original_text = best_match
                results.append({
                    'reference': ref,
                    'phrase': original_phrase,
                    'verse_text': original_text
                })
                found = True
        
        if not found:
            not_found.append(original_phrase)
    
    return results, not_found

def find_matching_files():
    """Find ALL matching TXT and XLSX files in the current directory"""
    files = os.listdir('.')
    
    # Find ALL TXT files that contain "body" in name
    txt_files = [f for f in files if f.lower().endswith('.txt') and 'body' in f.lower()]
    
    # Find ALL XLSX files
    xlsx_files = [f for f in files if f.lower().endswith(('.xlsx', '.xls'))]
    
    matches = []
    
    # Match files by their number prefix
    for txt_file in txt_files:
        # Extract the number prefix from TXT filename
        txt_number_match = re.search(r'(\d+)', txt_file)
        if txt_number_match:
            txt_number = txt_number_match.group(1)
            
            # Find matching XLSX file with same number
            for xlsx_file in xlsx_files:
                xlsx_number_match = re.search(r'(\d+)', xlsx_file)
                if xlsx_number_match and xlsx_number_match.group(1) == txt_number:
                    matches.append((txt_file, xlsx_file))
                    break
    
    # If no matches found by numbers, try simple extension-based matching
    if not matches:
        if txt_files and xlsx_files:
            matches.append((txt_files[0], xlsx_files[0]))
    
    return matches

def process_single_pair(txt_file, xlsx_file):
    """Process a single pair of files"""
    print(f"Processing: {txt_file} with {xlsx_file}")
    
    # Load data
    phrases = load_txt_phrases(txt_file)
    print(f"  Loaded {len(phrases)} phrases from TXT (after removing verse numbers)")
    
    verses, chapters = extract_xlsx_structure(xlsx_file)
    print(f"  Found {len(verses)} verses, and {len(chapters)} chapters in XLSX")
    
    # Show samples for verification
    print("  Sample phrases from TXT (after cleaning):")
    for i, phrase in enumerate(phrases[:3], 1):
        print(f"    {i}. '{phrase}'")
    
    print("  Sample verses from XLSX:")
    for i, (ref, text, row_num) in enumerate(verses[:3], 1):
        print(f"    {i}. {ref}: '{text[:50]}...'")
    
    # Match phrases to verses
    results, not_found = match_phrases_to_verses(phrases, verses)
    print(f"  Results: {len(results)} matched, {len(not_found)} not found")
    
    # Create output filename (same as TXT file but with .csv extension)
    output_file = os.path.splitext(txt_file)[0] + '.csv'
    
    # Write output
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter='\t')
        
        # Unmatched items first
        if not_found:
            writer.writerow(["ERROR: Could not find verses for these phrases:"])
            writer.writerow([])
            for phrase in not_found:
                writer.writerow([f"NOT_FOUND: {phrase}"])
            writer.writerow([])
            writer.writerow([])
        
        # Matches
        writer.writerow(["Reference", "Phrase"])
        for result in results:
            writer.writerow([result['reference'], result['phrase']])
    
    print(f"  Output saved to: {output_file}")
    
    # Show some examples
    if results:
        print("  First 3 matches:")
        for result in results[:3]:
            print(f"    {result['reference']} -> {result['phrase']}")
    
    # Debug: show some not found items
    if not_found:
        print("  Sample not found items:")
        for phrase in not_found[:5]:
            print(f"    NOT FOUND: '{phrase}'")
    
    return len(results), len(not_found)

def main():
    print("=== Bible Phrase to Verse Matcher ===")
    print("Finds the verse that contains each phrase")
    
    # Find ALL matching file pairs
    file_pairs = find_matching_files()
    
    if not file_pairs:
        print("No matching files found.")
        print("Looking for files like:")
        print("  - '01-Genesis body.txt'")
        print("  - '01-Genesis.xlsx'")
        print("Files in current directory:")
        for f in os.listdir('.'):
            print(f"  {f}")
        return
    
    print(f"Found {len(file_pairs)} matching file pair(s):")
    for txt, xlsx in file_pairs:
        print(f"  {txt} -> {xlsx}")
    
    print("\n" + "="*60)
    
    total_matched = 0
    total_not_found = 0
    
    # Process each file pair
    for txt_file, xlsx_file in file_pairs:
        try:
            matched, not_found = process_single_pair(txt_file, xlsx_file)
            total_matched += matched
            total_not_found += not_found
            print()
        except Exception as e:
            print(f"Error processing {txt_file}: {e}")
            import traceback
            traceback.print_exc()
            print()
    
    print("="*60)
    print("Processing complete!")
    print(f"Total across all files: {total_matched} matched, {total_not_found} not found")

if __name__ == "__main__":
    main()