import openpyxl
import fitz  # PyMuPDF
import logging
import sys
import os
import glob
from datetime import datetime
from collections import defaultdict

# ===================== LOGGING =====================
def setup_logging():
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f'logs/pdf_margin_annotator_{timestamp}.log'
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    logging.info(f"=== PDF Margin Annotator Started ===")
    logging.info(f"Log file: {log_filename}")
    return log_filename

# ===================== CONFIG =====================
reference_file = "margin_baseline_reference.txt"

# ===================== FILE PAIRS =====================
def find_pdf_excel_pairs():
    pdf_files = glob.glob("*.pdf")
    pairs = []
    
    for pdf_file in pdf_files:
        base_name = os.path.splitext(pdf_file)[0]
        excel_file = f"{base_name}.xlsx"
        if os.path.exists(excel_file):
            output_pdf = f"{base_name}_annotated.pdf"
            pairs.append({'pdf': pdf_file, 'excel': excel_file, 'output': output_pdf, 'base_name': base_name})
            logging.info(f"Found pair: {pdf_file} + {excel_file} -> {output_pdf}")
        else:
            logging.warning(f"PDF found but no matching Excel file: {pdf_file} (looking for {excel_file})")
    
    return pairs

# ===================== REFERENCE VALUES =====================
def load_reference_values(ref_file_path):
    logging.info(f"Loading reference values from: {ref_file_path}")
    ref_values = {}
    
    try:
        with open(ref_file_path, 'r') as f:
            lines = f.readlines()
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if ':' in line:
                key, value = line.split(':', 1)
                try:
                    ref_values[key.strip()] = float(value.strip())
                except ValueError:
                    logging.warning(f"Line {line_num}: Could not parse '{value.strip()}'")
        logging.info(f"Loaded {len(ref_values)} reference values")
        return ref_values
    except FileNotFoundError:
        logging.error(f"Reference file '{ref_file_path}' not found")
        sys.exit(1)

def get_reference_value(column_name, page_side, ref_values):
    column_mappings = {
        "Top Scripture Baseline Left (in)": "Top",
        "Top Scripture Baseline Right (in)": "Top",
        "Bottom Scripture Baseline Left (in)": "Bottom",
        "Bottom Scripture Baseline Right (in)": "Bottom",
        "Footnote Baseline (in)": "Footnote",
        "Book Intro Baseline (in)": "Book Intro",
        "Study Note Baseline (in)": "Study Note",
        "Article Baseline (in)": "Article",
        "Running Head Baseline (in)": "Running Head",
        "Page Number Baseline (in)": "Page Number",
        "Column 1 Max Width (in)": "Column 1 Max Width (in)",
        "Column 2 Max Width (in)": "Column 2 Max Width (in)",
        "Column Gap Width (in)": "Column Gap Width (in)",
        "Box Baseline (in)": "Box Baseline"
    }
    page_side_mappings = {
        "Column 1 Left Edge (in)": f"{page_side} Pages - Column 1 Left Edge (in)",
        "Column 1 Right Edge (in)": f"{page_side} Pages - Column 1 Right Edge (in)", 
        "Column 2 Left Edge (in)": f"{page_side} Pages - Column 2 Left Edge (in)",
        "Column 2 Right Edge (in)": f"{page_side} Pages - Column 2 Right Edge (in)"
    }
    
    if column_name in page_side_mappings:
        ref_key = page_side_mappings[column_name]
        return ref_values.get(ref_key, None)
    if column_name in column_mappings:
        ref_key = column_mappings[column_name]
        return ref_values.get(ref_key, None)
    return None

# ===================== MEASUREMENT TYPES =====================
def is_bottom_measurement(column_name):
    bottom_measurements = {
        "Bottom Scripture Baseline Left (in)",
        "Bottom Scripture Baseline Right (in)",
        "Bottom Scripture Baseline Column 1 (in)",
        "Bottom Scripture Baseline Column 2 (in)",
        "Footnote Baseline (in)",
        "Book Intro Baseline (in)",
        "Study Note Baseline (in)",
        "Article Baseline (in)",
        "Box Baseline (in)"
    }
    return column_name in bottom_measurements

def is_side_measurement(column_name):
    side_measurements = {
        "Column 1 Left Edge (in)",
        "Column 1 Right Edge (in)",
        "Column 2 Left Edge (in)", 
        "Column 2 Right Edge (in)",
        "Column 1 Max Width (in)",
        "Column 2 Max Width (in)",
        "Column Gap Width (in)"
    }
    return column_name in side_measurements

# ===================== COMMENT TEXT =====================
def create_comment_text(column_name, actual_value, reference_value, color_type):
    if is_bottom_measurement(column_name):
        from_position = "from the bottom"
    elif is_side_measurement(column_name):
        from_position = "from the side"
    else:
        from_position = "from the top"
    
    comment = f"{color_type} {column_name}. Text is {actual_value} inches {from_position}. Normally text is {reference_value}"
    return comment

# ===================== COLOR DETECTION =====================
COLOR_HEX_MAP = {
    "FFFF0000": "RED",
    "FFFFFF00": "YELLOW",
    "FF800080": "PURPLE",
    "FFFFA500": "ORANGE"
}

def get_cell_color(cell):
    try:
        if cell.fill.patternType != 'solid':
            return None
        rgb = cell.fill.start_color.rgb
        if rgb is None:
            return None
        return COLOR_HEX_MAP.get(rgb.upper())
    except:
        return None

# ===================== ANNOTATION COUNT VERIFICATION =====================
def add_verification_note(doc, expected_counts, actual_counts):
    """Add a verification note to the first page showing annotation count comparison"""
    if not doc.page_count:
        return
    
    # Create summary text
    summary_lines = []
    all_match = True
    
    for color in sorted(expected_counts.keys()):
        expected = expected_counts[color]
        actual = actual_counts.get(color, 0)
        
        if expected == actual:
            summary_lines.append(f"✓ {color}: {expected} expected, {actual} written")
        else:
            all_match = False
            summary_lines.append(f"✗ {color}: {expected} expected, {actual} written. {expected - actual} missing")
    
    if all_match:
        summary_text = "All annotations successfully written to PDF:\n" + "\n".join(summary_lines)
    else:
        summary_text = "Annotation count mismatch:\n" + "\n".join(summary_lines)
    
    # Add note to first page
    first_page = doc[0]
    page_width = first_page.rect.width
    page_height = first_page.rect.height
    
    # Position at top center of page
    position = (page_width / 2, 50)
    
    # Create the annotation
    annot = first_page.add_text_annot(position, summary_text)
    annot.set_info(title="Annotation Verification")
    annot.set_colors(stroke=[0, 0, 0], fill=[1, 1, 0.9])  # Light yellow background with black border
    annot.update()

# ===================== PROCESSING =====================
def process_file_pair(file_pair, ref_values):
    excel_file = file_pair['excel']
    pdf_file = file_pair['pdf']
    output_pdf = file_pair['output']
    base_name = file_pair['base_name']
    
    logging.info(f"=== Processing {base_name} ===")
    
    inch_to_pts = 72
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        logging.error(f"Error loading Excel {excel_file}: {e}")
        return False
    
    headers = {col: ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value}
    
    paged_comments = []
    expected_color_counts = defaultdict(int)
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            page_num = int(row[0].value)
            page_side = row[1].value
        except:
            continue
        
        for col_idx, cell in enumerate(row[2:], start=3):
            if cell.value is None or cell.value == "N/A":
                continue
            
            column_name = headers.get(col_idx, f"Column {col_idx}")
            color_type = get_cell_color(cell)
            if color_type is None:
                continue
            
            # Count expected annotations by color
            expected_color_counts[color_type] += 1
            
            actual_value = cell.value
            reference_value = get_reference_value(column_name, page_side, ref_values)
            
            if reference_value is not None:
                comment_text = create_comment_text(column_name, actual_value, reference_value, color_type)
            else:
                from_position = "from the bottom" if is_bottom_measurement(column_name) else "from the side" if is_side_measurement(column_name) else "from the top"
                comment_text = f"{color_type} {column_name}. Text is {actual_value} inches {from_position}. No reference available"
            
            paged_comments.append({
                "page_num": page_num,
                "y_inch": actual_value,
                "comment": comment_text,
                "color": color_type.lower(),
                "is_bottom": is_bottom_measurement(column_name)
            })
    
    logging.info(f"Total annotations prepared: {len(paged_comments)}")
    logging.info(f"Expected color counts: {dict(expected_color_counts)}")
    
    try:
        doc = fitz.open(pdf_file)
    except Exception as e:
        logging.error(f"Error opening PDF {pdf_file}: {e}")
        return False
    
    actual_color_counts = defaultdict(int)
    
    for entry in paged_comments:
        if entry["page_num"] < 1 or entry["page_num"] > len(doc):
            continue
        page = doc[entry["page_num"] - 1]
        page_height = page.rect.height
        
        x_pts = page.rect.width / 2
        if "Column 1" in entry["comment"]:
            x_pts = page.rect.width / 4
        elif "Column 2" in entry["comment"]:
            x_pts = 3 * page.rect.width / 4
        
        y_pts_pdf = page_height - (entry["y_inch"] * inch_to_pts) if entry["is_bottom"] else entry["y_inch"] * inch_to_pts
        y_pts_pdf = max(0, min(y_pts_pdf, page_height))
        
        annot = page.add_text_annot((x_pts, y_pts_pdf), entry["comment"])
        annot.set_info(title="Margin Check")
        
        # Count actual annotations by color
        actual_color_counts[entry["color"].upper()] += 1
        
        # Set colors
        if entry["color"] == "red":
            annot.set_colors(stroke=[1, 0, 0], fill=[1, 0.8, 0.8])
        elif entry["color"] == "yellow":
            annot.set_colors(stroke=[1, 1, 0], fill=[1, 1, 0.8])
        elif entry["color"] == "orange":
            annot.set_colors(stroke=[1, 0.6, 0], fill=[1, 0.9, 0.7])
        elif entry["color"] == "purple":
            annot.set_colors(stroke=[0.5, 0, 0.5], fill=[0.8, 0.7, 1])
        annot.update()
    
    # Add verification note to first page
    add_verification_note(doc, expected_color_counts, actual_color_counts)
    
    try:
        doc.save(output_pdf, incremental=False, garbage=4)
        doc.close()
        logging.info(f"Saved annotated PDF: {output_pdf}")
        logging.info(f"Actual color counts: {dict(actual_color_counts)}")
    except Exception as e:
        logging.error(f"Error saving PDF {output_pdf}: {e}")
        return False
    
    logging.info(f"=== {base_name} Complete: {len(paged_comments)} annotations added ===")
    return True

# ===================== MAIN =====================
def main():
    setup_logging()
    ref_values = load_reference_values(reference_file)
    file_pairs = find_pdf_excel_pairs()
    
    if not file_pairs:
        logging.error("No PDF/Excel pairs found!")
        sys.exit(1)
    
    total_processed, total_failed = 0, 0
    for pair in file_pairs:
        success = process_file_pair(pair, ref_values)
        if success:
            total_processed += 1
            print(f"✓ Successfully processed: {pair['base_name']}")
        else:
            total_failed += 1
            print(f"✗ Failed to process: {pair['base_name']}")
    
    print(f"\n=== BATCH COMPLETE ===")
    print(f"Successfully processed: {total_processed}")
    print(f"Failed: {total_failed}")
    if total_processed > 0:
        print(f"Annotated PDFs saved with '_annotated' suffix")

if __name__ == "__main__":
    main()