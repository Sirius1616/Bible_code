import openpyxl
import fitz  # PyMuPDF
import logging
import sys
import os
import glob
from datetime import datetime

# Configure logging
def setup_logging():
    """Set up detailed logging for troubleshooting."""
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f'logs/pdf_margin_annotator_{timestamp}.log'
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    logging.info(f"=== PDF Margin Annotator Started ===")
    logging.info(f"Log file: {log_filename}")
    return log_filename

# === CONFIG ===
reference_file = "margin_baseline_reference.txt"

def find_pdf_excel_pairs():
    """Find all PDF files and their corresponding Excel files in the current directory."""
    pdf_files = glob.glob("*.pdf")
    pairs = []
    
    for pdf_file in pdf_files:
        base_name = os.path.splitext(pdf_file)[0]
        excel_file = f"{base_name}.xlsx"
        
        if os.path.exists(excel_file):
            output_pdf = f"{base_name}_annotated.pdf"
            pairs.append({
                'pdf': pdf_file,
                'excel': excel_file,
                'output': output_pdf,
                'base_name': base_name
            })
            logging.info(f"Found pair: {pdf_file} + {excel_file} -> {output_pdf}")
        else:
            logging.warning(f"PDF found but no matching Excel file: {pdf_file} (looking for {excel_file})")
    
    return pairs

def load_reference_values(ref_file_path):
    """Load reference values from the margin baseline reference file."""
    logging.info(f"Loading reference values from: {ref_file_path}")
    ref_values = {}
    
    try:
        with open(ref_file_path, 'r') as f:
            lines = f.readlines()
        
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            if ':' in line:
                key, value = line.split(':', 1)
                try:
                    parsed_value = float(value.strip())
                    ref_values[key.strip()] = parsed_value
                    logging.debug(f"Line {line_num}: '{key.strip()}' = {parsed_value}")
                except ValueError:
                    logging.warning(f"Line {line_num}: Could not parse value '{value.strip()}' as float")
                    continue
                
        logging.info(f"Successfully loaded {len(ref_values)} reference values")
        return ref_values
        
    except FileNotFoundError:
        logging.error(f"Reference file '{ref_file_path}' not found")
        sys.exit(1)
    except Exception as e:
        logging.error(f"Error reading reference file: {e}")
        sys.exit(1)

def get_reference_value(column_name, page_side, ref_values):
    """Get the appropriate reference value based on column name and page side."""
    logging.debug(f"Getting reference value for column '{column_name}' on {page_side} page")
    
    # Define the mapping from Excel column names to reference keys
    column_mappings = {
        "Top Scripture Baseline Column 1 (in)": "Top",
        "Top Scripture Baseline Column 2 (in)": "Top",
        "Bottom Scripture Baseline Column 1 (in)": "Bottom",
        "Bottom Scripture Baseline Column 2 (in)": "Bottom",
        "Footnote Baseline (in)": "Footnote",
        "Book Intro Baseline (in)": "Book Intro",
        "Study Note Baseline (in)": "Study Note",
        "Article Baseline (in)": "Article",
        "Running Head Baseline (in)": "Running Head",
        "Page Number Baseline (in)": "Page Number",
        "Column 1 Max Width (in)": "Column 1 Max Width (in)",
        "Column 2 Max Width (in)": "Column 2 Max Width (in)",
        "Column Gap Width (in)": "Column Gap Width (in)",
        "Box Baseline (in)": "Box Baseline",
        "Subhead Baseline (in)": "Subhead Baseline (in)"
    }
    
    # Handle page-side dependent columns
    page_side_mappings = {
        "Column 1 Left Edge (in)": f"{page_side} Pages - Column 1 Left Edge (in)",
        "Column 1 Right Edge (in)": f"{page_side} Pages - Column 1 Right Edge (in)", 
        "Column 2 Left Edge (in)": f"{page_side} Pages - Column 2 Left Edge (in)",
        "Column 2 Right Edge (in)": f"{page_side} Pages - Column 2 Right Edge (in)"
    }
    
    # First try page-side dependent mapping
    if column_name in page_side_mappings:
        ref_key = page_side_mappings[column_name]
        logging.debug(f"Using page-side dependent mapping: '{column_name}' -> '{ref_key}'")
        ref_value = ref_values.get(ref_key)
        if ref_value is not None:
            logging.debug(f"Found reference value: {ref_value}")
            return ref_value
        else:
            logging.warning(f"Page-side dependent reference key '{ref_key}' not found in reference values")
    
    # Then try regular mapping
    if column_name in column_mappings:
        ref_key = column_mappings[column_name]
        logging.debug(f"Using regular mapping: '{column_name}' -> '{ref_key}'")
        ref_value = ref_values.get(ref_key)
        if ref_value is not None:
            logging.debug(f"Found reference value: {ref_value}")
            return ref_value
        else:
            logging.warning(f"Regular reference key '{ref_key}' not found in reference values")
    
    logging.warning(f"No mapping found for column '{column_name}'")
    return None

def is_bottom_measurement(column_name):
    """Check if a measurement is taken from the bottom of the page."""
    bottom_measurements = {
        "Bottom Scripture Baseline Column 1 (in)",
        "Bottom Scripture Baseline Column 2 (in)",
        "Footnote Baseline (in)",
        "Book Intro Baseline (in)",
        "Study Note Baseline (in)",
        "Article Baseline (in)",
        "Box Baseline (in)"
    }
    return column_name in bottom_measurements

def is_side_measurement(column_name):
    """Check if a measurement is taken from the side of the page."""
    side_measurements = {
        "Column 1 Left Edge (in)",
        "Column 1 Right Edge (in)",
        "Column 2 Left Edge (in)", 
        "Column 2 Right Edge (in)",
        "Column 1 Max Width (in)",
        "Column 2 Max Width (in)",
        "Column Gap Width (in)"
    }
    return column_name in side_measurements

def create_comment_text(column_name, actual_value, reference_value, color_type):
    """Create the comment text for the sticky note."""
    # Handle purple and orange specific text
    if color_type == "PURPLE":
        return "The two columns on this page do not align and they are not at their typical location."
    elif color_type == "ORANGE":
        return "This column is not aligned with the other column. The other column is in the correct position."
    
    # Original logic for red and yellow
    if is_bottom_measurement(column_name):
        from_position = "from the bottom"
    elif is_side_measurement(column_name):
        from_position = "from the side"
    else:
        from_position = "from the top"
    
    if reference_value is not None:
        comment = f"{color_type} {column_name}. Text is {actual_value} inches {from_position}. Normally text is {reference_value}"
    else:
        comment = f"{color_type} {column_name}. Text is {actual_value} inches {from_position}. No reference available"
    
    logging.debug(f"Created comment: '{comment}'")
    return comment

# === COLOR DETECTION ===
def is_red(cell):
    """Check if a cell is filled with red color."""
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            return cell.fill.fgColor.rgb.upper().endswith("FFC7CE")
        return False
    except:
        return False

def is_yellow(cell):
    """Check if a cell is filled with yellow color."""
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            return cell.fill.fgColor.rgb.upper().endswith("FFEB9C")
        return False
    except:
        return False

def is_purple(cell):
    """Check if a cell is filled with purple color."""
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            return cell.fill.fgColor.rgb.upper().endswith("E4DFEC")
        return False
    except:
        return False

def is_orange(cell):
    """Check if a cell is filled with orange color."""
    try:
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            return cell.fill.fgColor.rgb.upper().endswith("FDEADA")
        return False
    except:
        return False

def get_cell_color(cell):
    """Determine the color of a cell and return the color type as string."""
    if is_red(cell):
        return "RED"
    elif is_yellow(cell):
        return "YELLOW"
    elif is_purple(cell):
        return "PURPLE"
    elif is_orange(cell):
        return "ORANGE"
    return None

def add_summary_annotation(doc, color_counts, added_color_counts):
    """Add a summary annotation to the first page showing the comparison results."""
    try:
        summary_text = "ANNOTATION SUMMARY:\n\n"
        all_good = True
        missing_annotations = False
        
        for color in ["RED", "YELLOW", "PURPLE", "ORANGE"]:
            expected = color_counts[color]
            actual = added_color_counts[color]
            
            if expected == actual:
                summary_text += f"{color}: All {expected} annotations written successfully\n"
            else:
                all_good = False
                missing_annotations = True
                summary_text += f"{color}: {expected} expected, {actual} written. {expected - actual} missing\n"
        
        if all_good:
            summary_text = "All comments were successfully written into the PDF."
        elif missing_annotations:
            summary_text = "MISSING ANNOTATIONS:\n\n" + summary_text
        
        # Add summary to first page at top left
        first_page = doc[0]
        annot_rect = fitz.Rect(50, 50, 400, 200)  # Larger rectangle for summary
        annot = first_page.add_freetext_annot(annot_rect, summary_text)
        
        # Set font size and colors
        annot.set_info(title="Annotation Summary")
        annot.update()
        
        logging.info(f"Added summary annotation to first page: {summary_text}")
        return True
        
    except Exception as e:
        logging.error(f"Error adding summary annotation: {e}")
        return False

def process_file_pair(file_pair, ref_values):
    """Process a single PDF/Excel file pair."""
    excel_file = file_pair['excel']
    pdf_file = file_pair['pdf']
    output_pdf = file_pair['output']
    base_name = file_pair['base_name']
    
    logging.info(f"=== Processing {base_name} ===")
    logging.info(f"Excel: {excel_file}")
    logging.info(f"PDF: {pdf_file}")
    logging.info(f"Output: {output_pdf}")
    
    center_x_inch = ref_values.get("Bible Text Area Center Point (in)", 3.766)
    page_height_inches = ref_values.get("Page Height (in)", 10.5)
    logging.info(f"Using center point: {center_x_inch} inches")
    logging.info(f"Using page height: {page_height_inches} inches")
    
    inch_to_pts = 72
    
    # === LOAD EXCEL ===
    logging.info(f"Loading Excel file: {excel_file}")
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        logging.info(f"Excel file loaded successfully")
    except Exception as e:
        logging.error(f"Error loading Excel file {excel_file}: {e}")
        return False
    
    # Get column headers
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            headers[col] = cell_value
            logging.debug(f"Column {col}: '{cell_value}'")
    
    paged_comments = []
    annotations_found = 0
    color_counts = {"RED": 0, "YELLOW": 0, "PURPLE": 0, "ORANGE": 0}
    
    # Process each row
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            page_num = int(row[0].value)  # Column A = Page Number
            page_side = row[1].value      # Column B = Page Side (Left/Right)
        except (ValueError, TypeError):
            logging.debug(f"Row {row_num}: Invalid page number or side, skipping")
            continue
        
        logging.debug(f"Processing row {row_num}: Page {page_num} ({page_side})")
        
        # Check each column for colored cells
        for col_idx, cell in enumerate(row[2:], start=3):  # Start from column C
            if cell.value is None or cell.value == "N/A":
                continue
                
            column_name = headers.get(col_idx, f"Column {col_idx}")
            
            # Check if cell is colored
            color_type = get_cell_color(cell)
            if color_type:
                actual_value = cell.value
                color_counts[color_type] += 1
                
                logging.info(f"Found {color_type} cell: Page {page_num}, Column '{column_name}', Value: {actual_value}")
                
                # Get reference value
                reference_value = get_reference_value(column_name, page_side, ref_values)
                
                # Create comment text
                comment_text = create_comment_text(column_name, actual_value, reference_value, color_type)
                
                # Calculate Y position (handle bottom measurements)
                if is_bottom_measurement(column_name):
                    y_inch = actual_value
                    is_bottom_pos = True
                    logging.debug(f"Bottom measurement: {actual_value} inches from bottom")
                else:
                    y_inch = actual_value
                    is_bottom_pos = False
                    logging.debug(f"Top measurement: {actual_value} inches from top")
                
                paged_comments.append({
                    "page_num": page_num,
                    "y_inch": y_inch,
                    "comment": comment_text,
                    "color": color_type.lower(),
                    "is_bottom": is_bottom_pos,
                    "column_name": column_name
                })
                
                annotations_found += 1
                logging.info(f"Added annotation #{annotations_found}: {comment_text}")
    
    logging.info(f"Total annotations prepared for {base_name}: {annotations_found}")
    logging.info(f"Color counts: RED={color_counts['RED']}, YELLOW={color_counts['YELLOW']}, PURPLE={color_counts['PURPLE']}, ORANGE={color_counts['ORANGE']}")
    
    # === OPEN PDF AND ADD ANNOTATIONS ===
    logging.info(f"Opening PDF file: {pdf_file}")
    try:
        doc = fitz.open(pdf_file)
        logging.info(f"PDF opened successfully. Pages: {len(doc)}")
    except Exception as e:
        logging.error(f"Error opening PDF {pdf_file}: {e}")
        return False
    
    annotations_added = 0
    added_color_counts = {"RED": 0, "YELLOW": 0, "PURPLE": 0, "ORANGE": 0}
    
    # Process annotations
    for entry in paged_comments:
        if entry["page_num"] < 1 or entry["page_num"] > len(doc):
            logging.warning(f"Page {entry['page_num']} out of range, skipping")
            continue
            
        try:
            page = doc[entry["page_num"] - 1]
            page_height = page.rect.height
            
            # Calculate X position based on column reference
            x_pts = center_x_inch * inch_to_pts  # Default center position
            
            # Check if comment refers to Column 1 or Column 2 and adjust X position
            column_name = entry["column_name"]
            if "Column 1" in column_name:
                x_pts = (center_x_inch - 1.5) * inch_to_pts
                logging.info(f"Column 1 detected: Moving comment left 1.5 inches to {x_pts/inch_to_pts:.3f} inches")
            elif "Column 2" in column_name:
                x_pts = (center_x_inch + 1.5) * inch_to_pts
                logging.info(f"Column 2 detected: Moving comment right 1.5 inches to {x_pts/inch_to_pts:.3f} inches")
            else:
                logging.info(f"No column reference detected: Using center position {x_pts/inch_to_pts:.3f} inches")
            
            if entry.get("is_bottom", False):
                inches_from_bottom = entry["y_inch"]
                y_pts_pdf = page_height - (inches_from_bottom * inch_to_pts)
                logging.info(f"BOTTOM: Page {entry['page_num']}, {inches_from_bottom} inches from bottom = {y_pts_pdf} pts from top (page height: {page_height} pts)")
            else:
                y_pts = entry["y_inch"] * inch_to_pts
                y_pts_pdf = y_pts
                logging.info(f"TOP: Page {entry['page_num']}, {entry['y_inch']} inches from top = {y_pts_pdf} pts from top")
            
            # Ensure the annotation stays within page bounds
            if y_pts_pdf < 0:
                logging.warning(f"Annotation Y coordinate {y_pts_pdf} is above page, clamping to 0")
                y_pts_pdf = 0
            elif y_pts_pdf > page_height:
                logging.warning(f"Annotation Y coordinate {y_pts_pdf} is below page, clamping to {page_height}")
                y_pts_pdf = page_height
            
            annot = page.add_text_annot((x_pts, y_pts_pdf), entry["comment"])
            annot.set_info(title="Margin Check")
            
            # Set colors based on annotation type
            color = entry["color"]
            if color == "red":
                annot.set_colors(stroke=[1, 0, 0], fill=[1, 0.8, 0.8])
                added_color_counts["RED"] += 1
            elif color == "yellow":
                annot.set_colors(stroke=[1, 1, 0], fill=[1, 1, 0.8])
                added_color_counts["YELLOW"] += 1
            elif color == "purple":
                annot.set_colors(stroke=[0.5, 0, 0.5], fill=[0.8, 0.6, 0.8])
                added_color_counts["PURPLE"] += 1
            elif color == "orange":
                annot.set_colors(stroke=[1, 0.5, 0], fill=[1, 0.8, 0.6])
                added_color_counts["ORANGE"] += 1
                
            annot.update()
            annotations_added += 1
            logging.info(f"Added annotation to page {entry['page_num']} at ({x_pts}, {y_pts_pdf})")
            
        except Exception as e:
            logging.error(f"Error adding annotation to page {entry['page_num']}: {e}")
            continue
    
    # Add summary annotation showing comparison between expected and actual annotations
    add_summary_annotation(doc, color_counts, added_color_counts)
    
    # Save the PDF
    logging.info(f"Saving annotated PDF to: {output_pdf}")
    try:
        doc.save(output_pdf, incremental=False, garbage=4)
        doc.close()
        logging.info(f"PDF saved successfully")
    except Exception as e:
        logging.error(f"Error saving PDF {output_pdf}: {e}")
        return False
    
    logging.info(f"=== {base_name} Complete: {annotations_added} annotations added ===")
    logging.info(f"Added color counts: RED={added_color_counts['RED']}, YELLOW={added_color_counts['YELLOW']}, PURPLE={added_color_counts['PURPLE']}, ORANGE={added_color_counts['ORANGE']}")
    
    # Return success and the comparison results
    comparison_results = {
        "expected": color_counts,
        "actual": added_color_counts,
        "all_match": all(expected == added_color_counts[color] for color in color_counts)
    }
    
    return comparison_results

def main():
    """Main function."""
    log_file = setup_logging()
    
    # Load reference values (still needed for all files)
    ref_values = load_reference_values(reference_file)
    
    # Find all PDF/Excel pairs in the current directory
    file_pairs = find_pdf_excel_pairs()
    
    if not file_pairs:
        logging.error("No PDF/Excel file pairs found in current directory!")
        print("No PDF/Excel file pairs found!")
        print("Make sure you have matching PDF and XLSX files (same name, different extensions)")
        sys.exit(1)
    
    logging.info(f"Found {len(file_pairs)} file pair(s) to process")
    
    # Process each file pair
    total_processed = 0
    total_failed = 0
    comparison_results = []
    
    for file_pair in file_pairs:
        try:
            result = process_file_pair(file_pair, ref_values)
            if result:
                total_processed += 1
                comparison_results.append((file_pair['base_name'], result))
                print(f"✓ Successfully processed: {file_pair['base_name']}")
                
                # Print comparison results for this file
                print(f"  Comparison results:")
                for color in ["RED", "YELLOW", "PURPLE", "ORANGE"]:
                    expected = result["expected"][color]
                    actual = result["actual"][color]
                    status = "✓" if expected == actual else "✗"
                    print(f"    {color}: {expected} expected, {actual} written {status}")
                    
            else:
                total_failed += 1
                print(f"✗ Failed to process: {file_pair['base_name']}")
        except Exception as e:
            logging.error(f"Unexpected error processing {file_pair['base_name']}: {e}")
            total_failed += 1
            print(f"✗ Failed to process: {file_pair['base_name']} - {e}")
    
    # Final summary
    logging.info(f"=== BATCH PROCESSING COMPLETE ===")
    logging.info(f"Successfully processed: {total_processed}")
    logging.info(f"Failed: {total_failed}")
    
    print(f"\n=== BATCH PROCESSING COMPLETE ===")
    print(f"Successfully processed: {total_processed} files")
    print(f"Failed: {total_failed} files")
    
    # Print overall comparison summary
    if comparison_results:
        print(f"\n=== COMPARISON SUMMARY ===")
        all_files_good = True
        
        for file_name, result in comparison_results:
            file_status = "✓ ALL GOOD" if result["all_match"] else "✗ MISSING ANNOTATIONS"
            print(f"{file_name}: {file_status}")
            
            if not result["all_match"]:
                all_files_good = False
                for color in ["RED", "YELLOW", "PURPLE", "ORANGE"]:
                    expected = result["expected"][color]
                    actual = result["actual"][color]
                    if expected != actual:
                        print(f"  {color}: {expected} expected, {actual} written ({expected - actual} missing)")
        
        if all_files_good:
            print("All annotations were successfully written for all files!")
        else:
            print("Some annotations were missing from one or more files.")
    
    if total_processed > 0:
        print(f"\nAnnotated PDFs saved with '_annotated' suffix")

if __name__ == "__main__":
    main()