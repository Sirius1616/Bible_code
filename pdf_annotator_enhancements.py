# PDFWriter_extended.py

import openpyxl
import fitz  # PyMuPDF
import logging
import sys
import os
import glob
from datetime import datetime
from collections import defaultdict

# Configure logging
def setup_logging():
    if not os.path.exists('logs'):
        os.makedirs('logs')
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f'logs/pdf_margin_annotator_extended_{timestamp}.log'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename),
            logging.StreamHandler(sys.stdout)
        ]
    )
    logging.info("=== PDF Margin Annotator Extended Started ===")
    logging.info(f"Log file: {log_filename}")
    return log_filename

# === CONFIG ===
reference_file = "margin_baseline_reference.txt"

# === COLOR DETECTION ===
def get_cell_color(cell):
    """Detect cell color and return color type as string"""
    try:
        if cell.fill.fill_type != 'solid' or not cell.fill.start_color.rgb:
            return None
        
        rgb = cell.fill.start_color.rgb.upper()
        
        # Check for the specific RGB values found in your Excel file
        if rgb == "00FFC7CE":  # Red
            return "RED"
        elif rgb == "00FFEB9C":  # Yellow
            return "YELLOW"
        elif rgb == "00800080":  # Purple
            return "PURPLE"
        elif rgb == "00FFC000":  # Orange
            return "ORANGE"
        else:
            # Fallback to partial matching
            if "FFC7CE" in rgb or "C7CE" in rgb:  # Red
                return "RED"
            elif "FFEB9C" in rgb or "EB9C" in rgb:  # Yellow
                return "YELLOW"
            elif "800080" in rgb:  # Purple
                return "PURPLE"
            elif "FFC000" in rgb or "C000" in rgb:  # Orange
                return "ORANGE"
            else:
                logging.warning(f"Unknown color RGB: {rgb}")
                return None
    except Exception as e:
        logging.debug(f"Error detecting cell color: {e}")
        return None

# === LOAD REFERENCE VALUES ===
def load_reference_values(ref_file_path):
    logging.info(f"Loading reference values from: {ref_file_path}")
    ref_values = {}
    try:
        with open(ref_file_path, 'r') as f:
            for line in f:
                line = line.strip()
                if ':' in line:
                    key, value = line.split(':', 1)
                    try:
                        ref_values[key.strip()] = float(value.strip())
                    except ValueError:
                        continue
        logging.info(f"Loaded {len(ref_values)} reference values")
        return ref_values
    except FileNotFoundError:
        logging.error(f"Reference file '{ref_file_path}' not found")
        sys.exit(1)
    except Exception as e:
        logging.error(f"Error reading reference file: {e}")
        sys.exit(1)

# === FIND FILE PAIRS ===
def find_pdf_excel_pairs():
    pdf_files = glob.glob("*.pdf")
    pairs = []
    for pdf_file in pdf_files:
        # Skip already annotated files
        if "_annotated" in pdf_file:
            continue
            
        base_name = os.path.splitext(pdf_file)[0]
        excel_file = f"{base_name}.xlsx"
        if os.path.exists(excel_file):
            output_pdf = f"{base_name}_annotated.pdf"
            pairs.append({'pdf': pdf_file, 'excel': excel_file, 'output': output_pdf, 'base_name': base_name})
            logging.info(f"Found pair: {pdf_file} + {excel_file}")
        else:
            logging.warning(f"PDF found but no matching Excel file: {pdf_file}")
    return pairs

# === POSITION HELPERS ===
def is_bottom_measurement(column_name):
    bottom_measurements = {
        "Bottom Scripture Baseline Left (in)",
        "Bottom Scripture Baseline Right (in)",
        "Bottom Scripture Baseline Column 1 (in)",
        "Bottom Scripture Baseline Column 2 (in)",
        "Footnote Baseline (in)",
        "Book Intro Baseline (in)",
        "Study Note Baseline (in)",
        "Article Baseline (in)",
        "Box Baseline (in)"
    }
    return column_name in bottom_measurements

def is_side_measurement(column_name):
    side_measurements = {
        "Column 1 Left Edge (in)",
        "Column 1 Right Edge (in)",
        "Column 2 Left Edge (in)", 
        "Column 2 Right Edge (in)",
        "Column 1 Max Width (in)",
        "Column 2 Max Width (in)",
        "Column Gap Width (in)"
    }
    return column_name in side_measurements

def create_comment_text(column_name, actual_value, reference_value, color_type):
    # Special handling for purple and orange annotations
    if color_type == "PURPLE":
        return "The two columns on this page do not align and they are not at their typical location."
    elif color_type == "ORANGE":
        # Determine which column is orange and which is correct
        if "Column 1" in column_name:
            return "Column 1 is not aligned with Column 2. Column 2 is in the correct position."
        elif "Column 2" in column_name:
            return "Column 2 is not aligned with Column 1. Column 1 is in the correct position."
        else:
            return "This column is not aligned with the other column. The other column is in the correct position."
    
    # Standard handling for red and yellow annotations
    if is_bottom_measurement(column_name):
        from_position = "from the bottom"
    elif is_side_measurement(column_name):
        from_position = "from the side"
    else:
        from_position = "from the top"
    
    if reference_value is not None:
        comment = f"{color_type} {column_name}. Text is {actual_value} inches {from_position}. Normally text is {reference_value}"
    else:
        comment = f"{color_type} {column_name}. Text is {actual_value} inches {from_position}. No reference available"
    return comment

# === GET REFERENCE VALUE ===
def get_reference_value(column_name, page_side, ref_values):
    column_mappings = {
        "Top Scripture Baseline Left (in)": "Top",
        "Top Scripture Baseline Right (in)": "Top",
        "Bottom Scripture Baseline Left (in)": "Bottom",
        "Bottom Scripture Baseline Right (in)": "Bottom",
        "Footnote Baseline (in)": "Footnote",
        "Book Intro Baseline (in)": "Book Intro",
        "Study Note Baseline (in)": "Study Note",
        "Article Baseline (in)": "Article",
        "Running Head Baseline (in)": "Running Head",
        "Page Number Baseline (in)": "Page Number",
        "Column 1 Max Width (in)": "Column 1 Max Width (in)",
        "Column 2 Max Width (in)": "Column 2 Max Width (in)",
        "Column Gap Width (in)": "Column Gap Width (in)",
        "Box Baseline (in)": "Box Baseline"
    }
    page_side_mappings = {
        "Column 1 Left Edge (in)": f"{page_side} Pages - Column 1 Left Edge (in)",
        "Column 1 Right Edge (in)": f"{page_side} Pages - Column 1 Right Edge (in)", 
        "Column 2 Left Edge (in)": f"{page_side} Pages - Column 2 Left Edge (in)",
        "Column 2 Right Edge (in)": f"{page_side} Pages - Column 2 Right Edge (in)"
    }
    if column_name in page_side_mappings:
        return ref_values.get(page_side_mappings[column_name])
    return ref_values.get(column_mappings.get(column_name))

# === MAIN PROCESSING FUNCTION ===
def process_file_pair(file_pair, ref_values):
    excel_file = file_pair['excel']
    pdf_file = file_pair['pdf']
    output_pdf = file_pair['output']
    base_name = file_pair['base_name']
    
    logging.info(f"=== Processing {base_name} ===")
    
    inch_to_pts = 72
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except Exception as e:
        logging.error(f"Error loading Excel {excel_file}: {e}")
        return False
    
    headers = {col: ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value}
    
    paged_comments = []
    
    # Track color counts in Excel
    excel_color_counts = {"RED": 0, "YELLOW": 0, "ORANGE": 0, "PURPLE": 0}
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            page_num = int(row[0].value)
            page_side = row[1].value
        except:
            continue
        
        for col_idx, cell in enumerate(row[2:], start=3):
            if cell.value is None or cell.value == "N/A":
                continue
            
            column_name = headers.get(col_idx, f"Column {col_idx}")
            color_type = get_cell_color(cell)
            if color_type is None:
                continue
            
            # Increment Excel color count
            excel_color_counts[color_type] += 1
            
            actual_value = cell.value
            reference_value = get_reference_value(column_name, page_side, ref_values)
            
            comment_text = create_comment_text(column_name, actual_value, reference_value, color_type)
            
            paged_comments.append({
                "page_num": page_num,
                "y_inch": actual_value,
                "comment": comment_text,
                "color": color_type.lower(),
                "column_name": column_name,
                "is_bottom": is_bottom_measurement(column_name),
                "color_type": color_type  # Keep original color type for placement logic
            })
    
    logging.info(f"Total annotations prepared: {len(paged_comments)}")
    logging.info(f"Excel color counts: {excel_color_counts}")
    
    try:
        doc = fitz.open(pdf_file)
    except Exception as e:
        logging.error(f"Error opening PDF {pdf_file}: {e}")
        return False
    
    # Track actual written annotations
    pdf_color_counts = {"RED": 0, "YELLOW": 0, "ORANGE": 0, "PURPLE": 0}
    
    for entry in paged_comments:
        if entry["page_num"] < 1 or entry["page_num"] > len(doc):
            continue
            
        page = doc[entry["page_num"] - 1]
        page_height = page.rect.height
        
        # Horizontal placement - SPECIAL HANDLING FOR ORANGE
        if entry["color_type"] == "ORANGE":
            # Place over the specific column that is orange
            if "Column 1" in entry["column_name"]:
                x_pts = page.rect.width / 4  # Over Column 1
            elif "Column 2" in entry["column_name"]:
                x_pts = 3 * page.rect.width / 4  # Over Column 2
            else:
                x_pts = page.rect.width / 2  # Default center
        else:
            # Standard placement for other colors
            x_pts = page.rect.width / 2
            if "Column 1" in entry["column_name"]:
                x_pts = page.rect.width / 4
            elif "Column 2" in entry["column_name"]:
                x_pts = 3 * page.rect.width / 4
        
        # Vertical placement follows top/bottom rule
        y_pts_pdf = page_height - (entry["y_inch"] * inch_to_pts) if entry["is_bottom"] else entry["y_inch"] * inch_to_pts
        y_pts_pdf = max(0, min(y_pts_pdf, page_height))
        
        # Use text annotation instead of freetext to allow colors
        annot = page.add_text_annot((x_pts, y_pts_pdf), entry["comment"])
        annot.set_info(title="Margin Check")
        
        # Set colors
        if entry["color"] == "red":
            annot.set_colors(stroke=[1, 0, 0], fill=[1, 0.8, 0.8])
        elif entry["color"] == "yellow":
            annot.set_colors(stroke=[1, 1, 0], fill=[1, 1, 0.8])
        elif entry["color"] == "orange":
            annot.set_colors(stroke=[1, 0.6, 0], fill=[1, 0.9, 0.7])
        elif entry["color"] == "purple":
            annot.set_colors(stroke=[0.5, 0, 0.5], fill=[0.8, 0.7, 1])
        
        annot.update()
        pdf_color_counts[entry["color_type"]] += 1
    
    # ===================== ADD COMPARISON SUMMARY =====================
    first_page = doc[0]
    summary_lines = []
    all_matched = True
    
    for color in ["RED", "YELLOW", "ORANGE", "PURPLE"]:
        expected_count = excel_color_counts.get(color, 0)
        actual_count = pdf_color_counts.get(color, 0)
        
        if expected_count != actual_count:
            missing = expected_count - actual_count
            summary_lines.append(f"{expected_count} {color} annotations expected, {actual_count} written. {missing} missing.")
            all_matched = False
        elif expected_count > 0:
            summary_lines.append(f"✓ {color}: {expected_count} expected, {actual_count} written")
    
    if all_matched and any(excel_color_counts.values()):
        summary_lines.insert(0, "All annotations successfully written to the PDF.")
    elif not any(excel_color_counts.values()):
        summary_lines.append("No colored annotations found in Excel.")
    
    summary_text = "\n".join(summary_lines)
    
    # Add summary annotation to first page using text annotation (not freetext)
    annot = first_page.add_text_annot((50, 50), summary_text)
    annot.set_info(title="Annotation Summary")
    # For text annotations, we can set colors
    annot.set_colors(stroke=[0, 0, 0], fill=[1, 1, 0.9])  # Light yellow with black border
    annot.update()
    
    logging.info(f"PDF color counts: {pdf_color_counts}")
    
    try:
        doc.save(output_pdf, incremental=False, garbage=4)
        doc.close()
        logging.info(f"Saved annotated PDF: {output_pdf}")
    except Exception as e:
        logging.error(f"Error saving PDF {output_pdf}: {e}")
        return False
    
    logging.info(f"=== {base_name} Complete: {len(paged_comments)} annotations added ===")
    return True

# === MAIN ===
def main():
    setup_logging()
    ref_values = load_reference_values(reference_file)
    file_pairs = find_pdf_excel_pairs()
    
    if not file_pairs:
        logging.error("No PDF/Excel file pairs found!")
        sys.exit(1)

    success_count = 0
    for pair in file_pairs:
        if process_file_pair(pair, ref_values):
            success_count += 1
            print(f"✓ Successfully processed: {pair['base_name']}")
        else:
            print(f"✗ Failed to process: {pair['base_name']}")
    
    print(f"\n=== BATCH COMPLETE ===")
    print(f"Successfully processed: {success_count}")
    print(f"Failed: {len(file_pairs) - success_count}")
    if success_count > 0:
        print(f"Annotated PDFs saved with '_annotated' suffix")

if __name__ == "__main__":
    main()